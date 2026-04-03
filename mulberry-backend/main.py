# mulberry-backend/main.py
from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Dict, Any, List
import os
import json
import httpx
import logging

# 去除部分第三方库的啰嗦日志，保持终端洁净
logging.basicConfig(level=logging.INFO)

app = FastAPI(title="Mulberry Micro-framework API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ====== 系统路径物理常量定义 ======
# 获取当前 main.py 所在的绝对硬盘目录
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# 1. 角色配置文件地址存储库
CONFIG_FILE_PATH = os.path.join(BASE_DIR, "config.json")
# 2. 画布独立拓扑图数据存储库
CANVAS_DATA_FILE_PATH = os.path.join(BASE_DIR, "canvas_data.json")

# ====== 历史记录栈 (撤销/重做状态机) ======
history_stack: List[Dict[str, Any]] = []
current_index: int = -1

def get_current_status():
    return {"undo_count": max(0, current_index), "redo_count": max(0, len(history_stack) - 1 - current_index)}

@app.post("/api/history/reset")
async def reset_state(data: Dict[str, Any]):
    global history_stack, current_index
    history_stack = [data]
    current_index = 0
    return get_current_status()

@app.post("/api/history/push")
async def push_state(data: Dict[str, Any]):
    global history_stack, current_index
    if 0 <= current_index < len(history_stack) - 1:
        history_stack = history_stack[:current_index + 1]
    history_stack.append(data)
    current_index += 1
    return get_current_status()

@app.post("/api/history/undo")
async def undo_state():
    global current_index
    if current_index > 0: current_index -= 1
    return {"state": history_stack[current_index] if history_stack else None, "status": get_current_status()}

@app.post("/api/history/redo")
async def redo_state():
    global current_index
    if current_index < len(history_stack) - 1: current_index += 1
    return {"state": history_stack[current_index] if history_stack else None, "status": get_current_status()}


# ====== 图拓扑数据 (画布) 持久化读写接口 ======
@app.get("/api/canvas/data")
async def get_canvas_data():
    """只读接口：供前端在刷新或冷启动时调取上一次被主动保存的画布内容"""
    if os.path.exists(CANVAS_DATA_FILE_PATH):
        try:
            with open(CANVAS_DATA_FILE_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.warning(f"读取画布拓扑存量数据异常: {e}")
    # 如无文件，则返回规范的空列表约定
    return {"nodes": [], "edges": []}

@app.post("/api/canvas/data")
async def save_canvas_data(payload: Dict[str, Any]):
    """覆盖写入接口：接收前端手动的保存指令并物理落盘至独立的 .json 文件"""
    try:
        with open(CANVAS_DATA_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return {"status": "success"}
    except Exception as e:
        logging.error(f"图拓扑持久化硬盘写入失败: {e}")
        return {"status": "error", "message": str(e)}


# ====== 设置与校验 (Excel 路径与透明度配置) ======
@app.get("/api/settings/config")
async def get_config():
    if os.path.exists(CONFIG_FILE_PATH):
        try:
            with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return {
                    "path": data.get("excel_path", ""),
                    "upper_opacity": data.get("upper_opacity", 70), 
                    "node_opacity": data.get("node_opacity", 85)
                }
        except Exception as e:
            logging.warning(f"读取配置异常: {e}")
    return {"path": "", "upper_opacity": 70, "node_opacity": 85}

@app.post("/api/settings/config")
async def save_config(payload: Dict[str, Any]):
    data = {}
    if os.path.exists(CONFIG_FILE_PATH):
        try:
            with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            pass

    if "path" in payload:
        data["excel_path"] = payload.get("path", "").strip()
    if "upper_opacity" in payload:
        data["upper_opacity"] = payload.get("upper_opacity")
    if "node_opacity" in payload:
        data["node_opacity"] = payload.get("node_opacity")

    try:
        with open(CONFIG_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/settings/verify_excel")
async def verify_excel(payload: Dict[str, Any]):
    path = payload.get("path", "").strip()
    if not path: return {"status": "empty"}
    if not path.lower().endswith('.xlsx'): return {"status": "invalid"}
    if not os.path.exists(path): return {"status": "invalid"}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb.active
        headers = [str(cell).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if cell is not None]
        expected = ["角色名称", "调用的模型供应商", "调用的模型名称", "环境变量存储层级", "环境变量变量名"]
        if all(e in headers for e in expected): return {"status": "valid"}
        else: return {"status": "invalid"}
    except Exception:
        return {"status": "invalid"}

# ===================== 流式运行时执行图引擎 =====================

def find_node(nodes, node_id):
    return next((n for n in nodes if n['id'] == node_id), None)

async def _chat_generator(payload: Dict[str, Any]):
    # 提前初始化调用元数据，确保发生异常时也有默认值垫底
    provider, model_name, tokens_used = "未知", "未知", 0
    
    try:
        nodes = payload.get("nodes", [])
        edges = payload.get("edges", [])
        btn_id = payload.get("button_id")
        settings_path = payload.get("settings_path")

        # 容灾机制：若前端内存无数据，优先尝试退回硬盘读取最近固化地址
        if not settings_path or not os.path.exists(settings_path):
            if os.path.exists(CONFIG_FILE_PATH):
                try:
                    with open(CONFIG_FILE_PATH, 'r', encoding='utf-8') as f:
                        saved_config = json.load(f)
                        fallback_path = saved_config.get("excel_path", "")
                        if fallback_path and os.path.exists(fallback_path):
                            settings_path = fallback_path
                except Exception:
                    pass

        btn_in_edge = next((e for e in edges if e.get('target') == btn_id), None)
        if not btn_in_edge: raise ValueError("按钮缺失上游连线")
        
        current_content_node = find_node(nodes, btn_in_edge['source'])
        if current_content_node.get('data', {}).get('status') != '当前':
            raise ValueError("防呆拦截：当前上下文焦点状态并非处于『当前』")

        plan_edge = next((e for e in edges if e['source'] == current_content_node['id'] and find_node(nodes, e['target'])['type'] == 'contentNode'), None)
        if not plan_edge: raise ValueError("图有断层，找不到后续的计划接替节点")
        next_content_node = find_node(nodes, plan_edge['target'])

        curr_role_edge = next((e for e in edges if e['target'] == current_content_node['id'] and find_node(nodes, e['source'])['type'] == 'roleNode'), None)
        curr_role_name = find_node(nodes, curr_role_edge['source'])['data']['content'].strip()

        next_role_edge = next((e for e in edges if e['target'] == next_content_node['id'] and find_node(nodes, e['source'])['type'] == 'roleNode'), None)
        agent_role_node = find_node(nodes, next_role_edge['source'])
        agent_name = agent_role_node['data']['content'].strip()

        sys_prompt_edge = next((e for e in edges if e['target'] == agent_role_node['id'] and find_node(nodes, e['source'])['type'] == 'rolePromptNode'), None)
        sys_prompt = find_node(nodes, sys_prompt_edge['source'])['data']['content'] if sys_prompt_edge else ""

        meta_event = {
            "old_current_id": current_content_node['id'],
            "new_current_id": next_content_node['id']
        }
        yield f"event: meta\ndata: {json.dumps(meta_event)}\n\n"

        api_key = None
        model_found_flag = False
        
        if settings_path and os.path.exists(settings_path):
            import openpyxl
            wb = openpyxl.load_workbook(settings_path, read_only=True, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if str(row[0]).strip() == agent_name and agent_name != "用户":
                    provider = str(row[1]).strip() if row[1] else "未知"
                    model_name = str(row[2]).strip() if row[2] else "未知"
                    env_name = str(row[4]).strip() if row[4] else ""
                    api_key = os.getenv(env_name)
                    if api_key: model_found_flag = True
                    break

        user_input_text = current_content_node['data']['content']
        messages = []
        if sys_prompt: messages.append({"role": "system", "content": sys_prompt})
        messages.append({"role": "user", "content": f"{curr_role_name}：{user_input_text}"})

        if model_found_flag and provider.lower() in ['deepseek', 'openai']:
            base_url = "https://api.deepseek.com/chat/completions" if provider.lower() == 'deepseek' else "https://api.openai.com/v1/chat/completions"
            async with httpx.AsyncClient() as client:
                async with client.stream(
                    "POST", base_url,
                    headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                    json={"model": model_name, "messages": messages, "stream": True, "stream_options": {"include_usage": True}},
                    timeout=30.0
                ) as response:
                    async for chunk in response.aiter_lines():
                        if chunk.startswith("data: ") and chunk != "data: [DONE]":
                            try:
                                data_dict = json.loads(chunk[6:])
                                choices = data_dict.get("choices", [])
                                if choices and len(choices) > 0:
                                    delta = choices[0].get("delta", {}).get("content", "")
                                    if delta:
                                        yield f"event: text\ndata: {json.dumps({'text': delta})}\n\n"
                                
                                # 捕获最后带有 usage 指标的数据包 (支持 DeepSeek / OpenAI 的 include_usage 特性)
                                if "usage" in data_dict and data_dict["usage"]:
                                    tokens_used = data_dict["usage"].get("total_tokens", 0)
                            except Exception: pass
        else:
            fallback_text = f"你好，这里是本地演练流。我是角色『{agent_name}』。\n您的表格配置 { '成功命中角色' if provider != '未知' else '没有命中相关角色' }，但由于系统未找到此 API 请求密钥变量，已退回本地演示模式。\n\n看到了吗？这也是流式输出的一环！顺便，你的话“{user_input_text.strip()[:10]}...”收到了。"
            import asyncio
            for char in fallback_text:
                yield f"event: text\ndata: {json.dumps({'text': char})}\n\n"
                await asyncio.sleep(0.05)
            # 假装消耗了 Token
            tokens_used = len(fallback_text) + len(user_input_text)

        # 回传全部执行日志与消费信息给前端
        yield f"event: done\ndata: {json.dumps({'provider': provider, 'model': model_name, 'tokens': tokens_used})}\n\n"

    except Exception as e:
        # 当遇到断网、无环境等崩溃性异常时触发，连带模型元数据返回
        yield f"event: error\ndata: {json.dumps({'msg': str(e), 'provider': provider, 'model': model_name})}\n\n"

@app.post("/api/chat/run")
async def run_chat(request: Request):
    payload = await request.json()
    return StreamingResponse(_chat_generator(payload), media_type="text/event-stream")
